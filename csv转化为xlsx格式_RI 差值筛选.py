import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from openpyxl.styles import PatternFill
import string  # 用于生成字母序列


def process_file(file_path, output_folder, ri_threshold):
    """处理单个CSV文件，符合流程图逻辑"""
    global unprocessed_rows
    try:
        # 尝试使用 utf-8 编码读取
        df = pd.read_csv(file_path, encoding='utf-8')
    except UnicodeDecodeError:
        # 如果 utf-8 编码失败，则使用 gbk 编码重新读取
        df = pd.read_csv(file_path, encoding='gbk')

    # 确保列名中没有多余空格，避免匹配失败
    df.columns = df.columns.str.strip()

    # 填充“估计的浓度.”中的空值为 0，确保后续计算不会报错
    if '估计的浓度.' in df.columns:
        df['估计的浓度.'] = df['估计的浓度.'].fillna(0)
    else:
        messagebox.showwarning("警告", f"文件 {file_path} 中缺少 '估计的浓度.' 列，跳过处理")
        return

    # 提取 CAS 编号为 38818-55-2 的数据，并保留不处理
    if 'CAS 编号' in df.columns:
        unprocessed_rows = df[df['CAS 编号'].str.strip() == '38818-55-2']
        # 排除 38818-55-2 后的剩余数据
        df = df[df['CAS 编号'].str.strip() != '38818-55-2']

        # 如果存在未处理的 38818-55-2 行，则按照要求处理
        if not unprocessed_rows.empty:
            # 按“组分 RI”从小到大排序
            unprocessed_rows = unprocessed_rows.sort_values(by='组分 RI')

            # 生成字母序列 A, B, C, ...
            letters = list(string.ascii_uppercase)[:len(unprocessed_rows)]

            # 按顺序修改“用户定义的谱库化合物”列的值
            for idx, letter in enumerate(letters):
                unprocessed_rows.loc[unprocessed_rows.index[idx], '用户定义的谱库化合物'] = f"巨豆三烯酮{letter}"

    # 计算“组分 RI”和“谱库 RI”的差值，并创建新列“RI 差值”
    if '组分 RI' in df.columns and '谱库 RI' in df.columns:
        df['RI 差值'] = abs(df['组分 RI'] - df['谱库 RI'])
        # 根据用户输入的 RI 差值阈值过滤数据
        df = df[df['RI 差值'] <= ri_threshold]
    else:
        messagebox.showwarning("警告", f"文件 {file_path} 中缺少 '组分 RI' 或 '谱库 RI' 列，跳过处理")
        return

    # 查找并处理剩余的 "CAS 编号"
    unique_cas = df['CAS 编号'].unique()
    result_df = pd.DataFrame()
    highlight_rows = []  # 用于存储需要高亮的行索引

    for cas in unique_cas:
        # 显式复制 cas_group，避免 SettingWithCopyWarning
        cas_group = df[df['CAS 编号'] == cas].copy()

        if len(cas_group) > 1:
            # 找到差值最小的那一行
            min_diff_row = cas_group.loc[cas_group['RI 差值'].idxmin()].copy()
            # 将所有“估计的浓度.”相加后填入
            min_diff_row['估计的浓度.'] = cas_group['估计的浓度.'].sum()
            # 标记这行需要高亮
            highlight_rows.append(len(result_df))
            # 将处理后的行加入结果
            result_df = pd.concat([result_df, min_diff_row.to_frame().T], ignore_index=True)
        else:
            # 如果没有重复，直接加入结果
            result_df = pd.concat([result_df, cas_group], ignore_index=True)

    # 将未处理的 38818-55-2 数据追加到结果
    result_df = pd.concat([result_df, unprocessed_rows], ignore_index=True)

    # 保存结果为 XLSX 文件，保持原文件名，仅更改后缀
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    output_path = os.path.join(output_folder, f"转换后_{base_name}.xlsx")

    # 使用 ExcelWriter 保存并高亮特定行
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        result_df.to_excel(writer, index=False, sheet_name='结果')

        # 获取 openpyxl 的工作表
        workbook = writer.book
        worksheet = writer.sheets['结果']

        # 定义高亮样式（黄色填充）
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # 遍历需要高亮的行，并高亮“RI 差值”列
        for row_idx in highlight_rows:
            excel_row = row_idx + 2  # DataFrame 的行索引对应 Excel 的行，需加 2
            worksheet.cell(row=excel_row, column=result_df.columns.get_loc('RI 差值') + 1).fill = fill

    print(f"文件 {file_path} 处理完成，结果保存为 {output_path}")
def process_files(input_folder, output_folder, ri_threshold):
    """处理输入文件夹中的所有CSV文件"""
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    for file_name in os.listdir(input_folder):
        if file_name.endswith('.csv'):
            file_path = os.path.join(input_folder, file_name)
            process_file(file_path, output_folder, ri_threshold)

    print("所有文件处理完成。")
    messagebox.showinfo("完成", "所有文件已处理完成！")


# 图形界面的创建
def create_gui():
    def select_input_folder():
        folder = filedialog.askdirectory(title="选择输入文件夹")
        if folder:
            input_folder_var.set(folder)

    def select_output_folder():
        folder = filedialog.askdirectory(title="选择输出文件夹")
        if folder:
            output_folder_var.set(folder)

    def run_processing():
        input_folder = input_folder_var.get()
        output_folder = output_folder_var.get()

        if not os.path.exists(input_folder):
            messagebox.showerror("错误", "输入文件夹不存在，请重新选择！")
            return

        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # 弹出对话框，让用户输入 RI 差值阈值
        try:
            ri_threshold = float(simpledialog.askstring("输入 RI 差值阈值", "请输入用于过滤数据的 RI 差值阈值："))
        except ValueError:
            messagebox.showerror("错误", "请输入有效的数字")
            return

        process_files(input_folder, output_folder, ri_threshold)  # 调用批量处理函数

    # 创建窗口
    root = tk.Tk()
    root.title("CSV 转换与去重工具")
    root.geometry("500x300")

    # 输入文件夹选择
    tk.Label(root, text="选择输入文件夹:").pack(pady=10)
    input_folder_var = tk.StringVar()
    tk.Entry(root, textvariable=input_folder_var, width=50).pack(pady=5)
    tk.Button(root, text="选择文件夹", command=select_input_folder).pack(pady=5)

    # 输出文件夹选择
    tk.Label(root, text="选择输出文件夹:").pack(pady=10)
    output_folder_var = tk.StringVar()
    tk.Entry(root, textvariable=output_folder_var, width=50).pack(pady=5)
    tk.Button(root, text="选择文件夹", command=select_output_folder).pack(pady=5)

    # 运行按钮
    tk.Button(root, text="开始处理", command=run_processing, bg="green", fg="white").pack(pady=20)

    root.mainloop()


# 启动图形界面
if __name__ == "__main__":
    create_gui()
